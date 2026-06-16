from flask import Flask, render_template_string, jsonify, request, Response, make_response
from ultralytics import YOLO
import cv2
import threading
import json
import time
import os
import numpy as np
from datetime import datetime
import csv
from io import StringIO
import math
import urllib.parse 
import requests 
from requests.auth import HTTPDigestAuth, HTTPBasicAuth 
import re
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sqlite3 # <--- NOU: Adaugat pentru baza de date

# --- SETĂRI GLOBALE ---
SKIP_FRAMES = 1 
CONF_THRESH = 0.40 
CONFIG_FILE = "cameras.json"
HISTORY_FILE = "history.json"
DB_FILE = "traffic_events.db" # <--- NOU: Fisierul bazei de date
os.environ["OPENCV_FFMPEG_CAPTURE_OPTIONS"] = "rtsp_transport;tcp|timeout;2000"

app = Flask(__name__)

cameras_config = {}  
camera_data = {}     
history_data = {}    
thread_flags = {}    
latest_frames = {}   
system_logs = []     

TIME_SLOTS = [f"{h:02d}:00-{h+2:02d}:00" for h in range(0, 24, 2)]

# --- NOU: INITIALIZARE BAZA DE DATE ---
def init_db():
    with sqlite3.connect(DB_FILE) as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS events
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         cam_id TEXT,
                         event_type TEXT,
                         timestamp DATETIME)''')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_timestamp ON events(timestamp)')

def get_current_slot():
    now = datetime.now()
    start_h = (now.hour // 2) * 2
    return f"{start_h:02d}:00-{start_h+2:02d}:00"

def add_log(cam_name, event_type):
    now_str = datetime.now().strftime("%H:%M:%S | %d-%m-%Y")
    log_entry = {"time": now_str, "cam": cam_name, "event": event_type}
    system_logs.insert(0, log_entry) 
    if len(system_logs) > 1000: system_logs.pop()

def load_data():
    global cameras_config, history_data
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f: cameras_config = json.load(f)
        except: cameras_config = {}
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r') as f: history_data = json.load(f)
        except: history_data = {}

def save_config():
    with open(CONFIG_FILE, 'w') as f: json.dump(cameras_config, f, indent=4)

def save_history():
    with open(HISTORY_FILE, 'w') as f: json.dump(history_data, f, indent=4)

def init_camera_structures(cam_id):
    loc = cameras_config.get(cam_id, {}).get("location", "General")
    if cam_id not in camera_data:
        camera_data[cam_id] = {"in": 0, "out": 0, "status": "Initializare...", "name": cameras_config[cam_id].get("name", cam_id), "location": loc}
    today = datetime.now().strftime("%Y-%m-%d")
    if today not in history_data: history_data[today] = {}
    if cam_id not in history_data[today]:
        history_data[today][cam_id] = {slot: {"in": 0, "out": 0} for slot in TIME_SLOTS}

def ccw(A,B,C):
    return (C[1]-A[1]) * (B[0]-A[0]) > (B[1]-A[1]) * (C[0]-A[0])

def intersect(A,B,C,D):
    return ccw(A,C,D) != ccw(B,C,D) and ccw(A,B,C) != ccw(A,B,D)

def process_camera(cam_id):
    model = YOLO('yolo11n.pt') 
    cap = None
    
    while thread_flags.get(cam_id, False):
        config = cameras_config.get(cam_id)
        if not config: break
        
        user, password, ip = config['user'], config['password'], config['ip']
        port_http, port_rtsp = config.get('port_http', '80'), config.get('port_rtsp', '554')
        canal = config.get('channel', '1').strip()
        
        try: requests.get(f"http://{ip}:{port_http}/ISAPI/System/deviceInfo", auth=HTTPDigestAuth(user, password), timeout=2)
        except: pass
            
        safe_user, safe_pass = urllib.parse.quote(user), urllib.parse.quote(password)
        base_url = f"rtsp://{safe_user}:{safe_pass}@{ip}:{port_rtsp}"
        
        working_url = config.get('working_url')
        if not working_url:
            paths = [
                f"/Streaming/Channels/{canal}02", 
                f"/ISAPI/Streaming/channels/{canal}02", 
                f"/cam/realmonitor?channel={canal}&subtype=1"
            ]
            for p in paths:
                if not thread_flags.get(cam_id, False): break
                test_url = base_url + p
                test_cap = cv2.VideoCapture(test_url)
                if test_cap.isOpened():
                    ret, _ = test_cap.read()
                    if ret:
                        working_url = test_url
                        cameras_config[cam_id]['working_url'] = working_url
                        save_config()
                        test_cap.release()
                        break
                test_cap.release()

        if not working_url:
            if cam_id in camera_data: camera_data[cam_id]["status"] = "EROARE CONEXIUNE"
            time.sleep(5); continue

        target_w, target_h = 640, 360
        track_history = {} 
        counted_ids = []
        frame_count = 0
        cap = cv2.VideoCapture(working_url)
        
        while thread_flags.get(cam_id, False):
            ret, frame = cap.read()
            if not ret:
                if cam_id in camera_data: camera_data[cam_id]["status"] = "RECONECTARE..."
                cap.release(); time.sleep(2)
                if thread_flags.get(cam_id, False):
                    cap = cv2.VideoCapture(working_url)
                continue
                
            if cam_id in camera_data:
                camera_data[cam_id]["status"] = "ONLINE"
                camera_data[cam_id]["name"] = cameras_config[cam_id]["name"]
                camera_data[cam_id]["location"] = cameras_config[cam_id].get("location", "General")
                
            h_orig, w_orig = frame.shape[:2]
            scale_x, scale_y = target_w / w_orig, target_h / h_orig
            
            frame_small = cv2.resize(frame, (target_w, target_h))
            frame_disp = frame_small.copy()
            
            cur_cfg = cameras_config.get(cam_id, {})
            pt1, pt2 = cur_cfg.get('pt1'), cur_cfg.get('pt2')
            flip = cur_cfg.get('flip_dir', 1)

            L1, L2, nx, ny = None, None, 0, 0
            if pt1 and pt2 and pt1 != [0,0]:
                L1 = (int(pt1[0]), int(pt1[1]))
                L2 = (int(pt2[0]), int(pt2[1]))
                
                cv2.line(frame_disp, L1, L2, (0, 255, 0), 2)
                
                cx, cy = (L1[0]+L2[0])//2, (L1[1]+L2[1])//2
                dx, dy = L2[0]-L1[0], L2[1]-L1[1]
                mag = math.sqrt(dx**2 + dy**2) + 0.0001
                nx, ny = (-dy/mag)*flip, (dx/mag)*flip
                
                tx, ty = int(cx + nx*40), int(cy + ny*40)
                cv2.arrowedLine(frame_disp, (cx, cy), (tx, ty), (0, 255, 255), 3, tipLength=0.4)

            ret_jpg, buffer = cv2.imencode('.jpg', frame_disp)
            if ret_jpg: latest_frames[cam_id] = buffer.tobytes()
            
            frame_count += 1
            if frame_count % SKIP_FRAMES != 0: continue 
            
            results = model.track(frame_small, persist=True, verbose=False, imgsz=320, conf=CONF_THRESH, classes=[0], tracker="bytetrack.yaml")
            current_ids_in_frame = []

            if results and results[0].boxes and results[0].boxes.id is not None and L1 and L2:
                boxes = results[0].boxes.xyxy.cpu().numpy()
                ids = results[0].boxes.id.int().cpu().tolist()
                current_ids_in_frame = ids

                for box, track_id in zip(boxes, ids):
                    x1, y1, x2, y2 = map(int, box)
                    curr_point = ((x1 + x2) // 2, (y1 + y2) // 2)

                    if track_id in track_history and track_id not in counted_ids:
                        prev_point = track_history[track_id]
                        if intersect(prev_point, curr_point, L1, L2):
                            mx, my = curr_point[0] - prev_point[0], curr_point[1] - prev_point[1]
                            dot_product = (mx * nx) + (my * ny)
                            
                            current_slot = get_current_slot()
                            today = datetime.now().strftime("%Y-%m-%d")
                            now_db_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S") # NOU: Timp pentru DB
                            init_camera_structures(cam_id)
                            
                            if cam_id in camera_data:
                                if dot_product > 0:
                                    camera_data[cam_id]["in"] += 1
                                    history_data[today][cam_id][current_slot]["in"] += 1
                                    add_log(f"{cameras_config[cam_id]['name']} ({camera_data[cam_id]['location']})", "INTRARE") 
                                    # --- NOU: INSERARE IN BAZA DE DATE ---
                                    with sqlite3.connect(DB_FILE, timeout=10) as conn:
                                        conn.execute("INSERT INTO events (cam_id, event_type, timestamp) VALUES (?, 'IN', ?)", (cam_id, now_db_str))
                                else:
                                    camera_data[cam_id]["out"] += 1
                                    history_data[today][cam_id][current_slot]["out"] += 1
                                    add_log(f"{cameras_config[cam_id]['name']} ({camera_data[cam_id]['location']})", "IESIRE")
                                    # --- NOU: INSERARE IN BAZA DE DATE ---
                                    with sqlite3.connect(DB_FILE, timeout=10) as conn:
                                        conn.execute("INSERT INTO events (cam_id, event_type, timestamp) VALUES (?, 'OUT', ?)", (cam_id, now_db_str))
                                
                                counted_ids.append(track_id)
                                save_history() 

                    track_history[track_id] = curr_point
                    
            if frame_count % 100 == 0:
                track_history = {k: v for k, v in track_history.items() if k in current_ids_in_frame}
                
            time.sleep(0.01)
            
        if cap: cap.release()
    print(f"[THREAD] Stop Camera {cam_id}")

def start_camera_thread(cam_id):
    if cam_id not in thread_flags or not thread_flags[cam_id]:
        thread_flags[cam_id] = True
        t = threading.Thread(target=process_camera, args=(cam_id,))
        t.daemon = True; t.start()

def stop_camera_thread(cam_id):
    if cam_id in thread_flags: thread_flags[cam_id] = False

# --- RUTE API ---
@app.route('/data')
def data(): return jsonify({"live": camera_data, "logs": system_logs})

# --- NOU: RUTA PENTRU INTEROGAREA LA ORA EXACTA ---
@app.route('/api_exact_time', methods=['POST'])
def api_exact_time():
    req = request.json
    target_dt = req.get('datetime')
    if not target_dt: return jsonify({"success": False})
    
    # Includem toate secundele pana la sfarsitul minutului cerut
    target_dt += ":59" 
    
    result = {}
    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            # Numaram totalul de IN si OUT care au avut loc INAINTE sau EXACT la ora ceruta
            cursor.execute('''
                SELECT cam_id, 
                       SUM(CASE WHEN event_type = 'IN' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN event_type = 'OUT' THEN 1 ELSE 0 END)
                FROM events 
                WHERE timestamp <= ?
                GROUP BY cam_id
            ''', (target_dt,))
            for row in cursor.fetchall():
                result[row[0]] = {"in": row[1] or 0, "out": row[2] or 0}
    except Exception as e:
        print("Eroare DB:", e)
        return jsonify({"success": False})
        
    return jsonify({"success": True, "data": result})

@app.route('/api_history', methods=['POST'])
def api_history():
    req = request.json
    date_req = req.get('date', datetime.now().strftime("%Y-%m-%d"))
    if date_req not in history_data: return jsonify({"success": False, "data": {}})
    return jsonify({"success": True, "data": history_data[date_req]})

@app.route('/get_cameras')
def get_cameras(): return jsonify(cameras_config)

@app.route('/save_camera', methods=['POST'])
def save_camera():
    d = request.json
    cam_id = d.get('id')
    is_edit = cam_id in cameras_config
    if is_edit: stop_camera_thread(cam_id); time.sleep(1)
    
    cameras_config[cam_id] = {
        "name": d['name'], "location": d.get('location', 'General'), 
        "ip": d['ip'], "port_http": d['port_http'], "port_rtsp": d['port_rtsp'],
        "user": d['user'], "password": d['password'], "channel": d['channel'], 
        "pt1": cameras_config.get(cam_id, {}).get("pt1", [0,0]), 
        "pt2": cameras_config.get(cam_id, {}).get("pt2", [0,0]), 
        "flip_dir": cameras_config.get(cam_id, {}).get("flip_dir", 1), 
        "working_url": None
    }
    save_config(); init_camera_structures(cam_id); start_camera_thread(cam_id)
    return jsonify({"success": True})

@app.route('/save_line', methods=['POST'])
def save_line():
    d = request.json
    if d['id'] in cameras_config:
        cameras_config[d['id']].update({'pt1': d['pt1'], 'pt2': d['pt2'], 'flip_dir': d['flip_dir']})
        save_config(); return jsonify({"success": True})
    return jsonify({"success": False})

@app.route('/delete_camera', methods=['POST'])
def delete_camera():
    cid = request.json.get('id')
    if cid in cameras_config:
        stop_camera_thread(cid); time.sleep(0.5)
        del cameras_config[cid]; camera_data.pop(cid, None)
        save_config(); return jsonify({"success": True})
    return jsonify({"success": False})

@app.route('/fetch_channels', methods=['POST'])
def fetch_channels():
    d = request.json
    url_hik = f"http://{d['ip']}:{d['port_http']}/ISAPI/Streaming/channels"
    url_dah = f"http://{d['ip']}:{d['port_http']}/cgi-bin/configManager.cgi?action=getConfig&name=ChannelTitle"
    channels = []
    
    try:
        r = requests.get(url_hik, auth=HTTPDigestAuth(d['user'], d['password']), timeout=3)
        if r.status_code == 401: r = requests.get(url_hik, auth=HTTPBasicAuth(d['user'], d['password']), timeout=3)
        if r.status_code == 200:
            ids, names = re.findall(r'<id[^>]*>(.*?)</id>', r.text), re.findall(r'<channelName[^>]*>(.*?)</channelName>', r.text)
            for i in range(len(ids)):
                if str(ids[i]).endswith('01'): 
                    channels.append({"id": str(ids[i])[:-2], "name": names[i] if i < len(names) else f"Cam {str(ids[i])[:-2]}"})
            if channels: return jsonify({"success": True, "channels": channels})
    except: pass
    
    try:
        r = requests.get(url_dah, auth=HTTPDigestAuth(d['user'], d['password']), timeout=3)
        if r.status_code == 200:
            for line in r.text.split('\n'):
                if 'table.ChannelTitle[' in line and '].Name=' in line:
                    parts = line.split('=', 1)
                    if len(parts) == 2:
                        idx = re.search(r'\[(.*?)\]', line).group(1)
                        channels.append({"id": str(int(idx) + 1), "name": parts[1].strip()})
            if channels: return jsonify({"success": True, "channels": channels})
    except: pass
    
    return jsonify({"success": False, "error": "Eroare preluare liste."})

@app.route('/export_logs')
def export_logs():
    si = StringIO(); cw = csv.writer(si)
    cw.writerow(['Data', 'Sursa (Locatie)', 'Eveniment'])
    for log in system_logs: cw.writerow([log['time'], log['cam'], log['event']])
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=loguri_trafic.csv"
    output.headers["Content-type"] = "text/csv"
    return output

@app.route('/export_history')
def export_history():
    req_date = request.args.get('date', datetime.now().strftime("%Y-%m-%d"))
    req_cams = request.args.get('cams', 'ALL').split(',')
    req_slots = request.args.get('slots', 'ALL').split(',')
    
    if 'ALL' in req_slots: req_slots = TIME_SLOTS
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Raport {req_date}"
    
    header_fill = PatternFill(start_color="4facfe", end_color="4facfe", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    in_font = Font(color="008000", bold=True)
    out_font = Font(color="FF0000", bold=True)
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    headers = ['Locatie', 'Camera', 'Interval', 'IN', 'OUT', 'Persoane in Interior']
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    day_data = history_data.get(req_date, {})
    total_in = 0; total_out = 0; row_num = 2
    
    for cid, slots in day_data.items():
        if 'ALL' in req_cams or cid in req_cams:
            cam_info = cameras_config.get(cid, {})
            cam_name = cam_info.get("name", cid)
            cam_loc = cam_info.get("location", "General")
            
            for slot in req_slots:
                stats = slots.get(slot, {"in":0, "out":0})
                t_in = stats['in']; t_out = stats['out']; inside = t_in - t_out
                total_in += t_in; total_out += t_out
                
                ws.append([cam_loc, cam_name, slot, t_in, t_out, inside])
                
                for c_idx in range(1, 7):
                    cell = ws.cell(row=row_num, column=c_idx)
                    cell.alignment = center_align
                    cell.border = thin_border
                
                ws.cell(row=row_num, column=4).font = in_font
                ws.cell(row=row_num, column=5).font = out_font
                row_num += 1
                
    ws.append([])
    row_num += 1
    ws.append(["", "TOTAL GENERAL", "PENTRU FILTRELE ALESE", total_in, total_out, total_in - total_out])
    for c_idx in range(1, 7):
        cell = ws.cell(row=row_num, column=c_idx)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = total_fill
        cell.font = total_font
        if c_idx == 4: cell.font = Font(color="008000", bold=True)
        if c_idx == 5: cell.font = Font(color="FF0000", bold=True)
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 4

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output, 
        as_attachment=True, 
        download_name=f"Raport_Trafic_{req_date}.xlsx", 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def generate_frames(cam_id):
    while True:
        frame = latest_frames.get(cam_id)
        status = camera_data.get(cam_id, {}).get("status", "")
        if frame and status == "ONLINE":
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
        else:
            img = np.zeros((360, 640, 3), dtype=np.uint8)
            msg = "Asteptare conexiune..."
            if "conecteaza" in status.lower() or "Ping" in status: msg = "Se preia fluxul video..."
            elif "EROARE" in status: msg = "Eroare! Verifica IP-ul sau Portul."
            elif status == "RECONECTARE...": msg = "Reconectare in curs..."
            cv2.putText(img, msg, (80, 180), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
            _, buffer = cv2.imencode('.jpg', img)
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
        time.sleep(0.1)

@app.route('/video_feed/<cam_id>')
def video_feed(cam_id): return Response(generate_frames(cam_id), mimetype='multipart/x-mixed-replace; boundary=frame')

# --- WEB INTERFACE ---
HTML_PAGE = """
<!DOCTYPE html>
<html lang="ro">
<head>
    <meta charset="UTF-8">
    <title>Monitorizare Trafic</title>
    <style>
        body { font-family: sans-serif; background: #0a0a0a; color: #eee; margin: 0; padding: 20px; }
        h1 { text-align: center; color: #4facfe; margin-bottom: 20px; }
        .tabs { display: flex; gap: 10px; margin-bottom: 20px; justify-content: center; flex-wrap: wrap;}
        .tab-btn { background: #222; color: #fff; border: 1px solid #444; padding: 12px 25px; cursor: pointer; border-radius: 5px; font-weight: bold; transition:0.3s;}
        .tab-btn.active { background: #4facfe; color: #000; border-color: #4facfe; }
        .tab-content { display: none; }
        .tab-content.active { display: block; }

        .location-section { margin-bottom: 40px; background: #111; padding: 20px; border-radius: 12px; border: 1px solid #333; }
        .location-title { margin-top: 0; color: #ffb300; font-size: 1.5em; border-bottom: 2px solid #333; padding-bottom: 10px; margin-bottom: 20px; }
        .grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr)); gap: 20px;}
        .card { background: #1a1a1a; padding: 20px; border-radius: 12px; border-top: 5px solid #4facfe; box-shadow: 0 4px 15px rgba(0,0,0,0.5); display: flex; flex-direction: column;}
        
        .stats { display: flex; justify-content: space-around; margin-bottom: 15px; }
        .val { font-size: 2.5em; font-weight: bold; text-align: center;}
        .lbl { font-size: 0.8em; color: #aaa; text-align: center; }
        .in { color: #00e676; }
        .out { color: #ff1744; }
        .total-box { background: #222; border-radius: 10px; padding: 10px; text-align: center; margin-top: auto; margin-bottom: 15px; }
        .val-total { font-size: 2em; font-weight: bold; color: #4facfe; }
        
        .btn-play { background: #222; color: white; border: 1px solid #4facfe; padding: 10px; width: 100%; border-radius: 5px; cursor: pointer; font-weight: bold; transition: 0.2s; }
        .btn-play:hover { background: #4facfe; color:#000;}

        table { width: 100%; max-width: 1400px; margin: 0 auto; border-collapse: collapse; background: #161616; border-radius:10px; overflow:hidden;}
        th, td { padding: 12px; text-align: center; border-bottom: 1px solid #333; }
        th { background: #222; color: #4facfe; }
        
        .form-container { background: #161616; padding: 30px; border-radius: 12px; max-width: 650px; margin: 0 auto 30px auto; box-shadow: 0 4px 15px rgba(0,0,0,0.5);}
        input, select { width: 100%; padding: 12px; margin: 8px 0 20px 0; background: #222; color: #fff; border: 1px solid #444; border-radius: 6px; box-sizing: border-box; }
        .btn-submit { background: #00e676; color: #000; border: none; padding: 14px; width: 100%; font-weight: bold; cursor: pointer; border-radius: 6px; font-size: 1.1em; transition: 0.2s;}
        .btn-submit:hover { opacity: 0.8; }
        
        .modal-overlay { display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.95); z-index: 1000; text-align: center; }
        .modal-content { position: relative; display: inline-block; margin-top: 40px; background: #161616; padding: 20px; border-radius: 10px; border:2px solid #4facfe;}
        .video-box { position:relative; width:640px; height:360px; margin: 0 auto; background:#000; border-radius:5px; overflow:hidden; }
        canvas { position: absolute; top: 0; left: 0; cursor: crosshair; }
        
        .btn-modal { padding: 12px 25px; border: none; border-radius: 6px; font-weight: bold; cursor: pointer; margin: 15px 5px; font-size: 1em; }
        .btn-quick-pass { background: #333; color: white; border: 1px solid #555; padding: 5px 10px; border-radius: 4px; cursor: pointer; font-size: 0.85em; transition: 0.2s;}
        
        select[multiple] { height: 120px; outline:none; }
        select[multiple] option { padding: 5px; margin-bottom: 2px; }
        select[multiple] option:checked { background: #4facfe linear-gradient(0deg, #4facfe 0%, #4facfe 100%); color: #000; font-weight:bold;}
    </style>
</head>
<body>
    <h1>Sistem Monitorizare Trafic</h1>

    <div class="tabs">
        <button id="btn-live" class="tab-btn active" onclick="switchTab('live')">📊 Monitorizare Live</button>
        <button id="btn-exact" class="tab-btn" onclick="switchTab('exact')">⏱️ Căutare la Minut</button>
        <button id="btn-history" class="tab-btn" onclick="switchTab('history'); loadHistoryData();">📈 Căutare Istoric</button>
        <button id="btn-logs" class="tab-btn" onclick="switchTab('logs')">📋 Loguri Detaliate</button>
        <button id="btn-config" class="tab-btn" onclick="switchTab('config'); loadCams();">⚙️ Echipamente</button>
    </div>

    <div id="live" class="tab-content active"><div id="live-container"></div></div>

    <div id="exact" class="tab-content">
        <div class="form-container" style="max-width: 800px;">
            <h2 style="color:#ffb300; margin-top:0;">Câte persoane erau în interior la un moment dat?</h2>
            <label style="color:#aaa; font-weight:bold;">Selectează Data și Ora exactă:</label>
            <input type="datetime-local" id="exact-datetime-input" style="font-size: 1.2em;">
            <button class="btn-submit" onclick="fetchExactTime()" style="background:#ffb300;">🔍 Calculează Persoanele</button>
            
            <div id="exact-results" style="margin-top:30px; display:none;">
                <table style="width:100%;">
                    <thead>
                        <tr><th>Locație</th><th>Cameră</th><th>Au Intrat<br><span style="font-size:0.8em;color:#aaa;">(până la ora aleasă)</span></th><th>Au Ieșit<br><span style="font-size:0.8em;color:#aaa;">(până la ora aleasă)</span></th><th>Persoane În Interior</th></tr>
                    </thead>
                    <tbody id="exact-body"></tbody>
                </table>
            </div>
        </div>
    </div>

    <div id="history" class="tab-content">
        <div style="max-width: 1400px; margin: 0 auto;">
            <div style="background:#161616; padding:20px; border-radius:10px; border-left: 5px solid #ff9800; margin-bottom:20px; display:flex; gap:15px; align-items:flex-start; flex-wrap:wrap; box-shadow: 0 4px 10px rgba(0,0,0,0.3);">
                <div style="flex:1; min-width:150px;">
                    <label style="color:#aaa; font-weight:bold; font-size:0.9em; letter-spacing:1px;">ALEGE DATA</label>
                    <input type="date" id="hist-date" style="margin-top:8px; margin-bottom:0; font-size:1em;">
                </div>
                <div style="flex:1; min-width:180px;">
                    <label style="color:#aaa; font-weight:bold; font-size:0.9em; letter-spacing:1px;">LOCAȚII (Grup)</label>
                    <select id="hist-locs" multiple style="margin-top:8px; margin-bottom:0; font-size:1em;" onchange="filterCamsByLocation()">
                        <option value="ALL" selected>✅ Toate Locațiile</option>
                    </select>
                </div>
                <div style="flex:1; min-width:180px;">
                    <label style="color:#aaa; font-weight:bold; font-size:0.9em; letter-spacing:1px;">CAMERE <span style="font-weight:normal; font-size:0.8em;">(Ctrl+Click)</span></label>
                    <select id="hist-cam" multiple style="margin-top:8px; margin-bottom:0; font-size:1em;">
                        <option value="ALL" selected>✅ Arată Toate Camerele</option>
                    </select>
                </div>
                <div style="flex:1; min-width:180px;">
                    <label style="color:#aaa; font-weight:bold; font-size:0.9em; letter-spacing:1px;">INTERVALE <span style="font-weight:normal; font-size:0.8em;">(Ctrl+Click)</span></label>
                    <select id="hist-slots" multiple style="margin-top:8px; margin-bottom:0; font-size:1em;">
                        <option value="ALL" selected>✅ Arată Toate Intervalele</option>
                    </select>
                </div>
                <div style="display:flex; flex-direction:column; gap:10px; justify-content:flex-end; height:100%;">
                    <button class="btn-submit" style="width:auto; margin-bottom:0; padding:12px 25px; background:#4facfe;" onclick="loadHistoryData()">🔍 Afișează Sumarul</button>
                    <button class="btn-submit" style="width:auto; margin-bottom:0; padding:12px 25px; background:#ff9800;" onclick="exportHistory()">📥 Descarcă Raport</button>
                </div>
            </div>

            <div id="hist-summary" style="display:none; background:#222; padding:20px; border-radius:10px; margin-bottom:20px; text-align:center; border: 2px solid #4facfe; box-shadow: 0 4px 10px rgba(0,0,0,0.5);">
                <h3 style="margin:0; color:#aaa; font-weight:normal;">Sumar general pentru selecția curentă: <span id="sum-in" style="color:#00e676; font-weight:bold; margin-left:10px;">0 IN</span> <span style="color:#555; margin:0 10px;">|</span> <span id="sum-out" style="color:#ff1744; font-weight:bold;">0 OUT</span></h3>
                <h2 style="margin:10px 0 0 0; color:#fff; font-size:2em;">TOTAL PERSOANE ÎN INTERIOR: <span id="sum-inside" style="color:#ffb300; font-size:1.2em;">0</span></h2>
            </div>
            
            <div style="overflow-x: auto; background:#161616; border-radius:10px;">
                <table id="history-table"><thead id="history-head"></thead><tbody id="history-body"></tbody></table>
            </div>
        </div>
    </div>

    <div id="logs" class="tab-content">
        <div style="max-width: 1000px; margin: 0 auto;">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px;">
                <h2 style="color: #4facfe; margin:0;">Evenimente Detaliate</h2>
                <a href="/export_logs" style="background:#ff9800; padding:10px 20px; color:#000; font-weight:bold; text-decoration:none; border-radius:5px;">📥 Exportă Loguri</a>
            </div>
            <table><thead><tr><th>Ora Exactă</th><th>Nume Cameră</th><th>Tip Eveniment</th></tr></thead><tbody id="logs-body"></tbody></table>
        </div>
    </div>

    <div id="config" class="tab-content">
        <div class="form-container">
            <h3 style="color: #4facfe; margin-top:0;">Adaugă / Editează Sursă Video</h3>
            <label>ID Unic Intern</label><input type="text" id="f_id" placeholder="Ex: poarta1">
            <label>Nume Afișat în Aplicație</label><input type="text" id="f_name" placeholder="Ex: Intrare Principală">
            <label>Locație (Grupare)</label><input type="text" id="f_loc" placeholder="Ex: Hala 1, Clădirea B...">
            <label>Adresă IP</label><input type="text" id="f_ip" placeholder="82.76.164.107">
            <div style="display:flex; gap:10px;">
                <div style="flex:1;"><label>Port Web (HTTP)</label><input type="text" id="f_port_h" value="710"></div>
                <div style="flex:1;"><label>Port Video (RTSP)</label><input type="text" id="f_port_r" value="554"></div>
            </div>
            <label>Utilizator</label><input type="text" id="f_user" value="admin">
            <label>Parolă Sursă</label><input type="password" id="f_pass" value="Ifis_2022">
            <div style="margin-top:-15px; margin-bottom:15px; display:flex; gap:5px;">
                <button type="button" class="btn-quick-pass" onclick="document.getElementById('f_pass').value='Ifis_2022'">🔑 Ifis</button>
                <button type="button" class="btn-quick-pass" onclick="document.getElementById('f_pass').value='Asfa_2024'">🔑 Asfa</button>
            </div>
            
            <div style="display:flex; gap:10px; align-items:flex-end; margin-bottom:20px;">
                <div style="flex:1;"><label>Selectează Canalul</label><select id="f_ch" style="margin:0;"><option value="1">Canal 1</option></select></div>
                <button type="button" onclick="fetchCH()" style="height:45px; background:#ff9800; border:none; border-radius:6px; font-weight:bold; cursor:pointer; padding:0 15px; color:#000;">🔄 Scanează Canalele</button>
            </div>
            <button class="btn-submit" onclick="saveCam()">✅ Salvează Parametrii</button>
        </div>
        <table><thead><tr><th>Locație</th><th>Nume Sursă</th><th>Informații Reteaa</th><th>Status</th><th>Acțiuni</th></tr></thead><tbody id="cam-table"></tbody></table>
    </div>

    <div id="view-modal" class="modal-overlay">
        <div class="modal-content">
            <h2 id="view-modal-title" style="margin-top:0; color:#4facfe;">Flux Video Live</h2>
            <div class="video-box">
                <img id="view-stream-img" width="640" height="360" />
            </div>
            <div class="modal-buttons" style="display:block; text-align:center;">
                <button class="btn-modal" style="background:#ff1744; color:#fff;" onclick="closeViewModal()">❌ Închide Vizualizarea</button>
            </div>
        </div>
    </div>

    <div id="draw-modal" class="modal-overlay">
        <div class="modal-content">
            <h2 style="margin-top:0; color:#ffb300;">Definește Regula de Trecere</h2>
            <p style="color:#aaa; margin-top:0;">Trasează o linie nouă. Săgeata galbenă reprezintă direcția <b>INTRARE (IN)</b>.</p>
            <div class="video-box">
                <img id="modal-img" width="640" height="360">
                <canvas id="draw-canvas" width="640" height="360"></canvas>
            </div>
            <div class="modal-buttons">
                <button class="btn-modal" style="background:#00e676; color:#000;" onclick="saveL()">💾 Confirmă Regula</button>
                <button class="btn-modal" style="background:#4facfe; color:#000;" onclick="flipL()">🔄 Schimbă Sensul</button>
                <button class="btn-modal" style="background:#ff1744; color:#fff;" onclick="closeM()">❌ Anulează</button>
            </div>
        </div>
    </div>

    <script>
        const TIME_SLOTS_JS = ["00:00-02:00", "02:00-04:00", "04:00-06:00", "06:00-08:00", "08:00-10:00", "10:00-12:00", "12:00-14:00", "14:00-16:00", "16:00-18:00", "18:00-20:00", "20:00-22:00", "22:00-24:00"];
        let currentLiveDataCache = {};
        let cameraConfigCache = {}; 
        let activeDrawCamId = null;

        document.getElementById('hist-date').valueAsDate = new Date();
        const slotSelect = document.getElementById('hist-slots');
        TIME_SLOTS_JS.forEach(s => slotSelect.innerHTML += `<option value="${s}">${s}</option>`);

        let canvas = document.getElementById('draw-canvas');
        let ctx = canvas.getContext('2d');
        let isDrawing = false;
        let startX=0, startY=0, endX=0, endY=0;
        let flipDirection = 1; 

        const emptyImageSrc = "data:image/gif;base64,R0lGODlhAQABAAD/ACwAAAAAAQABAAACADs=";

        function switchTab(t) {
            document.querySelectorAll('.tab-content').forEach(e => e.classList.remove('active'));
            document.querySelectorAll('.tab-btn').forEach(e => e.classList.remove('active'));
            document.getElementById(t).classList.add('active'); event.target.classList.add('active');
        }

        function getMultiSelectValues(selectId) {
            return Array.from(document.getElementById(selectId).selectedOptions).map(opt => opt.value);
        }

        // --- NOU: FUNCTIA JAVASCRIPT PENTRU CAUTAREA EXACTA ---
        function fetchExactTime() {
            const dt = document.getElementById('exact-datetime-input').value;
            if(!dt) return alert("Alege data și ora!");
            
            // Format HTML5 datetime-local este YYYY-MM-DDTHH:MM, trebuie sa inlocuim T cu spatiu pentru SQLite
            const formattedDt = dt.replace('T', ' ');

            fetch('/api_exact_time', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({datetime: formattedDt})
            }).then(r=>r.json()).then(res => {
                if(!res.success) return alert("Eroare la interogarea bazei de date.");
                
                const tbody = document.getElementById('exact-body');
                tbody.innerHTML = '';
                let found = false;
                
                Object.keys(cameraConfigCache).forEach(cid => {
                    const data = res.data[cid];
                    if(data) {
                        found = true;
                        const inside = data.in - data.out;
                        const loc = cameraConfigCache[cid].location || 'General';
                        const name = cameraConfigCache[cid].name || cid;
                        
                        tbody.innerHTML += `<tr>
                            <td style="color:#ffb300; font-weight:bold;">${loc}</td>
                            <td style="color:#4facfe; font-weight:bold;">${name}</td>
                            <td style="color:#00e676;">${data.in}</td>
                            <td style="color:#ff1744;">${data.out}</td>
                            <td style="font-weight:bold; font-size:1.5em; color:#fff;">${inside}</td>
                        </tr>`;
                    }
                });
                
                if(!found) {
                    tbody.innerHTML = `<tr><td colspan="5" style="color:#ff1744; padding:20px;">Nu există evenimente înregistrate până la această dată/oră.</td></tr>`;
                }
                document.getElementById('exact-results').style.display = 'block';
            });
        }

        function filterCamsByLocation() {
            const locValues = getMultiSelectValues('hist-locs');
            const camSelect = document.getElementById('hist-cam');
            
            camSelect.innerHTML = '<option value="ALL" selected>✅ Arată Toate (din locațiile alese)</option>';
            
            Object.keys(cameraConfigCache).forEach(id => {
                const camLoc = cameraConfigCache[id].location || "General";
                if (locValues.includes("ALL") || locValues.includes(camLoc)) {
                    camSelect.innerHTML += `<option value="${id}">${cameraConfigCache[id].name}</option>`;
                }
            });
        }

        function loadHistoryData() {
            const d = document.getElementById('hist-date').value;
            const camValues = getMultiSelectValues('hist-cam');
            const slotValues = getMultiSelectValues('hist-slots');
            
            if(!d) return alert("Alege o dată validă!");
            if(camValues.length === 0) return alert("Selectează cel puțin o cameră!");
            if(slotValues.length === 0) return alert("Selectează cel puțin un interval!");
            
            fetch('/api_history', {
                method: 'POST', 
                headers:{'Content-Type':'application/json'},
                body: JSON.stringify({date: d})
            }).then(r=>r.json()).then(res => {
                const hHead = document.getElementById('history-head');
                const hBody = document.getElementById('history-body');
                const hSummary = document.getElementById('hist-summary');
                
                if(!res.success || Object.keys(res.data).length === 0) {
                    hHead.innerHTML = '';
                    hBody.innerHTML = `<tr><td colspan="100%" style="padding: 30px; color: #ff1744; font-size:1.2em;">Nu există istoric salvat pentru data de <b>${d}</b>.</td></tr>`;
                    hSummary.style.display = 'none';
                    return;
                }
                
                let allowedIds = [];
                if(camValues.includes("ALL")) {
                    Array.from(document.getElementById('hist-cam').options).forEach(opt => {
                        if(opt.value !== "ALL") allowedIds.push(opt.value);
                    });
                } else {
                    allowedIds = camValues;
                }

                let selectedSlots = slotValues.includes("ALL") ? TIME_SLOTS_JS : slotValues;
                hHead.innerHTML = `<tr><th>Locație</th><th style="min-width:150px;">Sursă Video</th>` + selectedSlots.map(s => `<th>Ora<br><span style="color:#fff; font-size:0.85em;">${s}</span></th>`).join('') + `</tr>`;
                
                let html = '';
                let grandTotalIn = 0; let grandTotalOut = 0;

                allowedIds.forEach(cid => {
                    if(!res.data[cid]) return; 
                    
                    let camName = cameraConfigCache[cid] ? cameraConfigCache[cid].name : cid;
                    let camLoc = cameraConfigCache[cid] ? cameraConfigCache[cid].location : "General";
                    
                    html += `<tr><td style="color:#ffb300; font-weight:bold;">${camLoc}</td><td style="font-weight:bold; color:#4facfe;">${camName}</td>`;
                    
                    selectedSlots.forEach(s => {
                        let slotData = res.data[cid][s] || {in:0, out:0};
                        grandTotalIn += slotData.in; grandTotalOut += slotData.out;
                        let inside = slotData.in - slotData.out;
                        
                        html += `<td>
                            <div style="margin-bottom:5px;">
                                <span style="color:#00e676; font-size:0.85em; font-weight:bold; margin-right:8px;">IN: ${slotData.in}</span>
                                <span style="color:#ff1744; font-size:0.85em; font-weight:bold;">OUT: ${slotData.out}</span>
                            </div>
                            <div style="background:#222; padding:3px; border-radius:4px; font-size:0.85em; border: 1px solid #444;">
                                Interior: <span style="color:#ffb300; font-weight:bold;">${inside}</span>
                            </div>
                        </td>`;
                    });
                    html += `</tr>`;
                });
                
                hBody.innerHTML = html;
                hSummary.style.display = 'block';
                document.getElementById('sum-in').innerText = grandTotalIn + " IN";
                document.getElementById('sum-out').innerText = grandTotalOut + " OUT";
                document.getElementById('sum-inside').innerText = (grandTotalIn - grandTotalOut);
            });
        }

        function exportHistory() {
            const d = document.getElementById('hist-date').value;
            const cams = getMultiSelectValues('hist-cam').join(',');
            const slots = getMultiSelectValues('hist-slots').join(',');
            if(!d) return alert("Alege o dată pentru a exporta!");
            window.location.href = `/export_history?date=${d}&cams=${encodeURIComponent(cams)}&slots=${encodeURIComponent(slots)}`;
        }

        function fetchCH() {
            const d = {ip: document.getElementById('f_ip').value, port_http: document.getElementById('f_port_h').value, user: document.getElementById('f_user').value, password: document.getElementById('f_pass').value};
            fetch('/fetch_channels', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(d)}).then(r=>r.json()).then(res=>{
                if(res.success) {
                    const s = document.getElementById('f_ch'); s.innerHTML = '';
                    res.channels.forEach(c => s.innerHTML += `<option value="${c.id}">[Canal ${c.id}] ${c.name}</option>`); alert("Date preluate cu succes!");
                } else alert("A eșuat interogarea echipamentului.");
            });
        }

        function saveCam() {
            const d = {id:document.getElementById('f_id').value, name:document.getElementById('f_name').value, location:document.getElementById('f_loc').value, ip:document.getElementById('f_ip').value, port_http:document.getElementById('f_port_h').value, port_rtsp:document.getElementById('f_port_r').value, user:document.getElementById('f_user').value, password:document.getElementById('f_pass').value, channel:document.getElementById('f_ch').value};
            fetch('/save_camera', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(d)}).then(() => {alert("Echipament salvat!"); loadCams(); document.getElementById('f_id').value='';});
        }

        function loadCams() {
            fetch('/get_cameras').then(r=>r.json()).then(data => {
                cameraConfigCache = data; 
                const b = document.getElementById('cam-table'); b.innerHTML = '';
                
                const hLocSelect = document.getElementById('hist-locs');
                hLocSelect.innerHTML = '<option value="ALL" selected>✅ Toate Locațiile</option>';
                let unice = new Set();
                
                Object.keys(data).forEach(id => {
                    let loc = data[id].location || "General";
                    unice.add(loc);
                    b.innerHTML += `<tr><td style="color:#ffb300; font-weight:bold;">${loc}</td><td>${data[id].name}</td><td style="font-size:0.8em; color:#aaa;">H:${data[id].port_http} / R:${data[id].port_rtsp}</td><td id="stat-${id}">...</td><td><button onclick="openM('${id}')" style="background:#ffb300; border:none; padding:5px 10px; border-radius:4px; font-weight:bold; cursor:pointer;">📏 Reguli</button> <button onclick="delCam('${id}')" style="background:#ff1744; color:white; border:none; padding:5px 10px; border-radius:4px; font-weight:bold; cursor:pointer;">Șterge</button></td></tr>`;
                });
                
                unice.forEach(l => hLocSelect.innerHTML += `<option value="${l}">${l}</option>`);
                filterCamsByLocation(); 
            });
        }

        function openM(id) {
            activeID = id; flipDirection = 1; 
            startX = 0; startY = 0; endX = 0; endY = 0; 
            ctx.clearRect(0,0,640,360);
            document.getElementById('draw-modal').style.display='block'; 
            document.getElementById('modal-img').src=`/video_feed/${id}?t=${Date.now()}`;
        }

        canvas.onmousedown = (e) => { const r = canvas.getBoundingClientRect(); startX = e.clientX-r.left; startY = e.clientY-r.top; isDrawing = true; };
        canvas.onmousemove = (e) => { if(!isDrawing) return; const r = canvas.getBoundingClientRect(); endX = e.clientX-r.left; endY = e.clientY-r.top; drawUI(); };
        canvas.onmouseup = () => isDrawing = false;

        function drawUI() {
            ctx.clearRect(0,0,640,360); ctx.beginPath(); ctx.moveTo(startX, startY); ctx.lineTo(endX, endY); ctx.strokeStyle='#00e676'; ctx.lineWidth=4; ctx.stroke();
            let cx=(startX+endX)/2, cy=(startY+endY)/2, dx=endX-startX, dy=endY-startY, mag=Math.sqrt(dx*dx+dy*dy)+0.001, nx=(-dy/mag)*flipDirection, ny=(dx/mag)*flipDirection;
            ctx.beginPath(); ctx.moveTo(cx, cy); ctx.lineTo(cx+nx*40, cy+ny*40); ctx.strokeStyle='#ff0'; ctx.lineWidth=3; ctx.stroke();
            ctx.fillStyle='#ff0'; ctx.beginPath(); ctx.arc(cx+nx*40, cy+ny*40, 6, 0, Math.PI*2); ctx.fill();
        }
        
        function flipL() { flipDirection *= -1; drawUI(); }
        function saveL() { fetch('/save_line', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({id:activeID, pt1:[startX,startY], pt2:[endX,endY], flip_dir:flipDirection})}).then(()=>closeM()); }
        function closeM() { document.getElementById('draw-modal').style.display='none'; document.getElementById('modal-img').src=emptyImageSrc; }
        function delCam(id) { if(confirm("Anulezi preluarea pentru această cameră?")) fetch('/delete_camera', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({id:id})}).then(()=>loadCams()); }

        function openViewModal(id) {
            document.getElementById('view-modal').style.display = 'block';
            document.getElementById('view-stream-img').src = `/video_feed/${id}?t=${Date.now()}`;
            document.getElementById('view-modal-title').innerText = "Live: " + (currentLiveDataCache[id] ? currentLiveDataCache[id].name : id);
        }
        function closeViewModal() {
            document.getElementById('view-modal').style.display = 'none';
            document.getElementById('view-stream-img').src = emptyImageSrc;
        }

        loadCams();

        setInterval(() => {
            fetch('/data').then(r=>r.json()).then(d => {
                currentLiveDataCache = d.live;
                const container = document.getElementById('live-container');
                
                Object.keys(d.live).forEach(id => {
                    const info = d.live[id];
                    let locName = info.location || "General";
                    let safeLocId = "loc-" + locName.replace(/\W/g, ''); 
                    
                    let locSection = document.getElementById(safeLocId);
                    if(!locSection) {
                        locSection = document.createElement('div');
                        locSection.id = safeLocId;
                        locSection.className = 'location-section';
                        locSection.innerHTML = `<h2 class="location-title">📍 Locație: ${locName}</h2><div class="grid" id="grid-${safeLocId}"></div>`;
                        container.appendChild(locSection);
                    }
                    
                    const g = document.getElementById(`grid-${safeLocId}`);
                    let c = document.getElementById(`card-${id}`);
                    let s_ui = info.status === "ONLINE" ? "🟢 FUNCȚIONEAZĂ" : (info.status.includes("Conectare")||info.status.includes("Ping") ? "🟡 "+info.status : "🔴 "+info.status);
                    let borderColor = info.status === "ONLINE" ? "#00e676" : "#ff1744";

                    if(!c) {
                        c = document.createElement('div'); c.id = `card-${id}`; c.className = 'card'; g.appendChild(c);
                        c.innerHTML = `
                            <div style="display:flex; justify-content:space-between; margin-bottom:15px;"><span style="font-weight:bold; font-size:1.2em;" id="n-${id}">${info.name}</span><span id="s-${id}" style="font-size:0.8em; color:#ddd; font-weight:bold;">${s_ui}</span></div>
                            <div class="stats"><div><div class="val in" id="in-${id}">${info.in}</div><div class="lbl">S-AU CONTORIZAT IN</div></div><div><div class="val out" id="out-${id}">${info.out}</div><div class="lbl">S-AU CONTORIZAT OUT</div></div></div>
                            <div class="total-box"><div class="val-total" id="tot-${id}" style="color:#4facfe;">${info.in - info.out}</div><div style="font-size:0.8em; text-transform:uppercase;">Persoane interior astazi</div></div>
                            <button class="btn-play" onclick="openViewModal('${id}')">🎥 Afișează Camera (Live)</button>
                        `;
                    } else {
                        if(c.parentElement !== g) g.appendChild(c); 
                        
                        document.getElementById(`in-${id}`).innerText = info.in; document.getElementById(`out-${id}`).innerText = info.out;
                        document.getElementById(`tot-${id}`).innerText = info.in - info.out; document.getElementById(`s-${id}`).innerText = s_ui;
                        c.style.borderTopColor = borderColor;
                    }
                    if(document.getElementById(`stat-${id}`)) document.getElementById(`stat-${id}`).innerText = s_ui;
                });
                
                Array.from(document.querySelectorAll('.card')).forEach(card => {
                    let cid = card.id.replace('card-','');
                    if(!d.live[cid]) card.remove();
                });
                Array.from(document.querySelectorAll('.location-section')).forEach(sec => {
                    if(sec.querySelector('.grid').children.length === 0) sec.remove();
                });
                
                const b = document.getElementById('logs-body'); b.innerHTML='';
                if(d.logs) d.logs.forEach(l => b.innerHTML += `<tr><td>${l.time}</td><td>${l.cam}</td><td style="color:${l.event==='INTRARE'?'#00e676':'#ff1744'}; font-weight:bold;">${l.event}</td></tr>`);
            });
        }, 1000);
    </script>
</body>
</html>
"""

@app.route('/')
def index(): return render_template_string(HTML_PAGE)

if __name__ == '__main__':
    init_db() 
    load_data()
    for cid in cameras_config: init_camera_structures(cid); start_camera_thread(cid)
    app.run(host='0.0.0.0', port=5000)
