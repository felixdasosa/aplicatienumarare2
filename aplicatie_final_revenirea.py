from flask import Flask, render_template_string, jsonify, request, Response, make_response, send_file
import cv2
import threading
import json
import time
import os
import numpy as np
from datetime import datetime
import csv
from io import BytesIO, StringIO
import urllib.parse 
import requests 
from requests.auth import HTTPDigestAuth, HTTPBasicAuth 
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- SETĂRI GLOBALE ---
CONFIG_FILE = "cameras.json"
HISTORY_FILE = "history.json"
os.environ["OPENCV_FFMPEG_CAPTURE_OPTIONS"] = "rtsp_transport;tcp|timeout;2000"

app = Flask(__name__)

cameras_config = {}  
camera_data = {}     
history_data = {}    
thread_flags = {}    
latest_frames_jpeg = {}  
system_logs = []     

TIME_SLOTS = [f"{h:02d}:00-{h+2:02d}:00" for h in range(0, 24, 2)]

def get_current_slot():
    now = datetime.now()
    start_h = (now.hour // 2) * 2
    return f"{start_h:02d}:00-{start_h+2:02d}:00"

def add_log(cam_name, event_type):
    now_str = datetime.now().strftime("%H:%M:%S | %d-%m-%Y")
    system_logs.insert(0, {"time": now_str, "cam": cam_name, "event": event_type}) 
    if len(system_logs) > 1000: system_logs.pop()

def load_data():
    global cameras_config, history_data
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f: cameras_config = json.load(f)
        except: cameras_config = {}
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r') as f: history_data = json.load(f)
        except: history_data = {}

def save_config():
    with open(CONFIG_FILE, 'w') as f: json.dump(cameras_config, f, indent=4)

def save_history():
    with open(HISTORY_FILE, 'w') as f: json.dump(history_data, f, indent=4)

def init_camera_structures(cam_id):
    loc = cameras_config.get(cam_id, {}).get("location", "General")
    if cam_id not in camera_data:
        camera_data[cam_id] = {"in": 0, "out": 0, "status": "Initializare...", "name": cameras_config[cam_id].get("name", cam_id), "location": loc}
    today = datetime.now().strftime("%Y-%m-%d")
    if today not in history_data: history_data[today] = {}
    if cam_id not in history_data[today]:
        history_data[today][cam_id] = {slot: {"in": 0, "out": 0} for slot in TIME_SLOTS}


# =========================================================================
# THREAD 1: ASCULTATOR ALERTE HIKVISION (Numărare IN/OUT)
# =========================================================================
def hikvision_alert_listener(cam_id):
    config = cameras_config.get(cam_id)
    if not config: return
    
    ip, user, password = config['ip'], config['user'], config['password']
    port_http = config.get('port_http', '80')
    target_channel = str(config.get('channel', '1'))
    
    url = f"http://{ip}:{port_http}/ISAPI/Event/notification/alertStream"
    
    while thread_flags.get(cam_id, False):
        if cam_id in camera_data and not str(camera_data[cam_id]["status"]).startswith("EROARE:"):
            camera_data[cam_id]["status"] = "Se conecteaza..."

        try:
            response = requests.get(url, auth=HTTPDigestAuth(user, password), stream=True, timeout=15)
            if response.status_code == 401:
                response = requests.get(url, auth=HTTPBasicAuth(user, password), stream=True, timeout=15)
                
            if response.status_code == 200:
                camera_data[cam_id]["status"] = "ONLINE"
                event_block = ""
                
                for line in response.iter_lines():
                    if not thread_flags.get(cam_id, False): break
                    if line:
                        decoded_line = line.decode('utf-8', errors='ignore')
                        event_block += decoded_line + "\n"
                        
                        if "</EventNotificationAlert>" in decoded_line:
                            if "<eventType>linedetection</eventType>" in event_block.lower() and f"<channelID>{target_channel}</channelID>" in event_block:
                                sens_calculat = "IN"
                                match_dir = re.search(r'<direction>(.*?)</direction>', event_block, re.IGNORECASE)
                                if match_dir:
                                    dir_val = match_dir.group(1).lower()
                                    if dir_val in ['atob', 'forward', 'lefttoright', 'left-right']: sens_calculat = "IN"
                                    elif dir_val in ['btoa', 'backward', 'righttoleft', 'right-left']: sens_calculat = "OUT"
                                
                                current_slot = get_current_slot()
                                today = datetime.now().strftime("%Y-%m-%d")
                                init_camera_structures(cam_id)
                                
                                if sens_calculat == "IN":
                                    camera_data[cam_id]["in"] += 1
                                    history_data[today][cam_id][current_slot]["in"] += 1
                                    add_log(f"{cameras_config[cam_id]['name']} ({camera_data[cam_id]['location']})", "INTRARE")
                                else:
                                    camera_data[cam_id]["out"] += 1
                                    history_data[today][cam_id][current_slot]["out"] += 1
                                    add_log(f"{cameras_config[cam_id]['name']} ({camera_data[cam_id]['location']})", "IESIRE")
                                save_history()
                            event_block = ""

            elif response.status_code == 401:
                camera_data[cam_id]["status"] = "EROARE: Parola sau Utilizator gresit!"
                time.sleep(10)
            elif response.status_code == 404:
                camera_data[cam_id]["status"] = "EROARE: Port HTTP gresit sau ISAPI dezactivat!"
                time.sleep(5)
            else:
                camera_data[cam_id]["status"] = f"EROARE: Raspuns Server {response.status_code}"
                time.sleep(5)
                
        except requests.exceptions.Timeout:
            camera_data[cam_id]["status"] = "EROARE: Timp expirat (15s). Verifica IP-ul si Portul HTTP!"
            time.sleep(5)
        except requests.exceptions.ConnectionError:
            camera_data[cam_id]["status"] = "EROARE: Conexiune refuzata. Camera offline sau IP invalid!"
            time.sleep(5)
        except Exception:
            camera_data[cam_id]["status"] = "EROARE: Conexiune HTTP intrerupta!"
            time.sleep(5)


# =========================================================================
# THREAD 2: PRELUARE VIDEO HIBRIDĂ (Cea mai sigură metodă)
# =========================================================================
def video_reader(cam_id):
    """Încearcă să ia poza prin HTTP. Dacă e invalidă (XML), trece pe RTSP."""
    config = cameras_config.get(cam_id)
    if not config: return
    
    ip, user, password = config['ip'], config['user'], config['password']
    port_http = config.get('port_http', '80')
    port_rtsp = config.get('port_rtsp', '554')
    canal = config.get('channel', '1').strip()
    
    # URL-urile de test
    urls_http = [
        f"http://{ip}:{port_http}/ISAPI/Streaming/channels/{canal}01/picture",
        f"http://{ip}:{port_http}/ISAPI/Streaming/channels/{canal}02/picture"
    ]
    safe_user, safe_pass = urllib.parse.quote(user), urllib.parse.quote(password)
    url_rtsp = f"rtsp://{safe_user}:{safe_pass}@{ip}:{port_rtsp}/Streaming/Channels/{canal}02"
    
    mod_video = "HTTP" # Incepem cu varianta usoara
    cap = None
    
    while thread_flags.get(cam_id, False):
        if "EROARE:" in str(camera_data.get(cam_id, {}).get("status", "")):
            time.sleep(1)
            continue
            
        if mod_video == "HTTP":
            cadru_valid = False
            for url in urls_http:
                try:
                    response = requests.get(url, auth=HTTPDigestAuth(user, password), timeout=3)
                    if response.status_code == 401:
                        response = requests.get(url, auth=HTTPBasicAuth(user, password), timeout=3)
                        
                    if response.status_code == 200:
                        # CRITIC: Validăm cu OpenCV ca să nu servim XML browserului
                        image_array = np.asarray(bytearray(response.content), dtype=np.uint8)
                        img = cv2.imdecode(image_array, cv2.IMREAD_COLOR)
                        
                        if img is not None:
                            img = cv2.resize(img, (640, 360))
                            ret_jpg, buffer = cv2.imencode('.jpg', img, [cv2.IMWRITE_JPEG_QUALITY, 60])
                            if ret_jpg:
                                latest_frames_jpeg[cam_id] = buffer.tobytes()
                                cadru_valid = True
                                break # Iesim din bucla, url-ul a mers
                except Exception:
                    pass
            
            if not cadru_valid:
                # Nu primim poza valabilă, trecem la planul B (RTSP)
                mod_video = "RTSP"
            else:
                time.sleep(0.3) # Pauză pentru HTTP ca să nu blocăm rețeaua
                
        elif mod_video == "RTSP":
            if cap is None or not cap.isOpened():
                cap = cv2.VideoCapture(url_rtsp, cv2.CAP_FFMPEG)
                
            ret, frame = cap.read()
            if ret:
                frame = cv2.resize(frame, (640, 360))
                ret_jpg, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 60])
                if ret_jpg:
                    latest_frames_jpeg[cam_id] = buffer.tobytes()
            else:
                cap.release()
                cap = None
                time.sleep(2)
                
            time.sleep(0.05)
            
    if cap: cap.release()


def start_camera_threads(cam_id):
    if cam_id not in thread_flags or not thread_flags[cam_id]:
        thread_flags[cam_id] = True
        threading.Thread(target=hikvision_alert_listener, args=(cam_id,), daemon=True).start()
        threading.Thread(target=video_reader, args=(cam_id,), daemon=True).start()

def stop_camera_thread(cam_id):
    if cam_id in thread_flags: thread_flags[cam_id] = False


# =========================================================================
# RUTE API ȘI GENERARE FLUX VIDEO SIGUR
# =========================================================================
@app.route('/data')
def data(): return jsonify({"live": camera_data, "logs": system_logs})

@app.route('/api_history', methods=['POST'])
def api_history():
    req = request.json
    date_req = req.get('date', datetime.now().strftime("%Y-%m-%d"))
    if date_req not in history_data: return jsonify({"success": False, "data": {}})
    return jsonify({"success": True, "data": history_data[date_req]})

@app.route('/get_cameras')
def get_cameras(): return jsonify(cameras_config)

@app.route('/save_camera', methods=['POST'])
def save_camera():
    d = request.json
    cam_id = d.get('id')
    is_edit = cam_id in cameras_config
    if is_edit: stop_camera_thread(cam_id); time.sleep(1)
    
    cameras_config[cam_id] = {
        "name": d['name'], "location": d.get('location', 'General'), 
        "ip": d['ip'], "port_http": d['port_http'], "port_rtsp": d['port_rtsp'],
        "user": d['user'], "password": d['password'], "channel": d['channel']
    }
    save_config(); init_camera_structures(cam_id); start_camera_threads(cam_id)
    return jsonify({"success": True})

@app.route('/delete_camera', methods=['POST'])
def delete_camera():
    cid = request.json.get('id')
    if cid in cameras_config:
        stop_camera_thread(cid); time.sleep(0.5)
        del cameras_config[cid]; camera_data.pop(cid, None)
        save_config(); return jsonify({"success": True})
    return jsonify({"success": False})

@app.route('/fetch_channels', methods=['POST'])
def fetch_channels():
    d = request.json
    url_hik = f"http://{d['ip']}:{d['port_http']}/ISAPI/Streaming/channels"
    channels = []
    try:
        r = requests.get(url_hik, auth=HTTPDigestAuth(d['user'], d['password']), timeout=3)
        if r.status_code == 401: r = requests.get(url_hik, auth=HTTPBasicAuth(d['user'], d['password']), timeout=3)
        if r.status_code == 200:
            ids, names = re.findall(r'<id[^>]*>(.*?)</id>', r.text), re.findall(r'<channelName[^>]*>(.*?)</channelName>', r.text)
            for i in range(len(ids)):
                if str(ids[i]).endswith('01'): 
                    channels.append({"id": str(ids[i])[:-2], "name": names[i] if i < len(names) else f"Cam {str(ids[i])[:-2]}"})
            if channels: return jsonify({"success": True, "channels": channels})
    except: pass
    return jsonify({"success": False, "error": "Eroare preluare liste."})

@app.route('/export_logs')
def export_logs():
    si = StringIO(); cw = csv.writer(si)
    cw.writerow(['Data', 'Sursa (Locatie)', 'Eveniment'])
    for log in system_logs: cw.writerow([log['time'], log['cam'], log['event']])
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=loguri_trafic.csv"
    output.headers["Content-type"] = "text/csv"
    return output

@app.route('/export_history')
def export_history():
    req_date = request.args.get('date', datetime.now().strftime("%Y-%m-%d"))
    req_cams = request.args.get('cams', 'ALL').split(',')
    req_slots = request.args.get('slots', 'ALL').split(',')
    
    if 'ALL' in req_slots: req_slots = TIME_SLOTS
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Raport {req_date}"
    
    header_fill = PatternFill(start_color="4facfe", end_color="4facfe", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    in_font = Font(color="008000", bold=True)
    out_font = Font(color="FF0000", bold=True)
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    headers = ['Locatie', 'Camera', 'Interval', 'IN', 'OUT', 'Persoane in Interior']
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    day_data = history_data.get(req_date, {})
    total_in, total_out, row_num = 0, 0, 2
    
    for cid, slots in day_data.items():
        if 'ALL' in req_cams or cid in req_cams:
            cam_info = cameras_config.get(cid, {})
            cam_name = cam_info.get("name", cid)
            cam_loc = cam_info.get("location", "General")
            
            for slot in req_slots:
                stats = slots.get(slot, {"in":0, "out":0})
                t_in, t_out = stats['in'], stats['out']
                inside = t_in - t_out
                total_in += t_in
                total_out += t_out
                
                ws.append([cam_loc, cam_name, slot, t_in, t_out, inside])
                for c_idx in range(1, 7):
                    cell = ws.cell(row=row_num, column=c_idx)
                    cell.alignment = center_align
                    cell.border = thin_border
                ws.cell(row=row_num, column=4).font = in_font
                ws.cell(row=row_num, column=5).font = out_font
                row_num += 1
                
    ws.append([])
    row_num += 1
    ws.append(["", "TOTAL GENERAL", "PENTRU FILTRELE ALESE", total_in, total_out, total_in - total_out])
    for c_idx in range(1, 7):
        cell = ws.cell(row=row_num, column=c_idx)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = total_fill
        cell.font = total_font
        if c_idx == 4: cell.font = Font(color="008000", bold=True)
        if c_idx == 5: cell.font = Font(color="FF0000", bold=True)
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 4

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name=f"Raport_Trafic_{req_date}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def generate_frames(cam_id):
    """Preia datele valide și previne crash-urile (imaginea ruptă)."""
    while True:
        frame = latest_frames_jpeg.get(cam_id)
        status = str(camera_data.get(cam_id, {}).get("status", ""))
        
        if frame and status == "ONLINE":
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
        else:
            img = np.zeros((360, 640, 3), dtype=np.uint8)
            # Fără diacritice, OpenCV nu afișează corect caracterele speciale
            safe_status = status.replace("ă", "a").replace("ț", "t").replace("ș", "s").replace("î", "i").replace("â", "a")
            
            if "EROARE:" in safe_status: 
                cv2.putText(img, "EROARE:", (20, 160), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                cv2.putText(img, safe_status[7:67], (20, 200), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1)
            else:
                cv2.putText(img, "Asteptare flux video...", (80, 180), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
                
            _, buffer = cv2.imencode('.jpg', img)
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
            
        time.sleep(0.1)

@app.route('/video_feed/<cam_id>')
def video_feed(cam_id): return Response(generate_frames(cam_id), mimetype='multipart/x-mixed-replace; boundary=frame')

# =========================================================================
# WEB INTERFACE (HTML)
# =========================================================================
HTML_PAGE = """
<!DOCTYPE html>
<html lang="ro">
<head>
    <meta charset="UTF-8">
    <title>Monitorizare Trafic (Edge AI)</title>
    <style>
        body { font-family: sans-serif; background: #0a0a0a; color: #eee; margin: 0; padding: 20px; }
        h1 { text-align: center; color: #4facfe; margin-bottom: 20px; }
        .tabs { display: flex; gap: 10px; margin-bottom: 20px; justify-content: center; flex-wrap: wrap;}
        .tab-btn { background: #222; color: #fff; border: 1px solid #444; padding: 12px 25px; cursor: pointer; border-radius: 5px; font-weight: bold; transition:0.3s;}
        .tab-btn.active { background: #4facfe; color: #000; border-color: #4facfe; }
        .tab-content { display: none; }
        .tab-content.active { display: block; }

        .location-section { margin-bottom: 40px; background: #111; padding: 20px; border-radius: 12px; border: 1px solid #333; }
        .location-title { margin-top: 0; color: #ffb300; font-size: 1.5em; border-bottom: 2px solid #333; padding-bottom: 10px; margin-bottom: 20px; }
        .grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr)); gap: 20px;}
        .card { background: #1a1a1a; padding: 20px; border-radius: 12px; border-top: 5px solid #4facfe; box-shadow: 0 4px 15px rgba(0,0,0,0.5); display: flex; flex-direction: column;}
        
        .stats { display: flex; justify-content: space-around; margin-bottom: 15px; }
        .val { font-size: 2.5em; font-weight: bold; text-align: center;}
        .lbl { font-size: 0.8em; color: #aaa; text-align: center; }
        .in { color: #00e676; }
        .out { color: #ff1744; }
        .total-box { background: #222; border-radius: 10px; padding: 10px; text-align: center; margin-top: auto; margin-bottom: 15px; }
        .val-total { font-size: 2em; font-weight: bold; color: #4facfe; }
        
        .btn-play { background: #222; color: white; border: 1px solid #4facfe; padding: 10px; width: 100%; border-radius: 5px; cursor: pointer; font-weight: bold; transition: 0.2s; }
        .btn-play:hover { background: #4facfe; color:#000;}

        table { width: 100%; max-width: 1400px; margin: 0 auto; border-collapse: collapse; background: #161616; border-radius:10px; overflow:hidden;}
        th, td { padding: 12px; text-align: center; border-bottom: 1px solid #333; }
        th { background: #222; color: #4facfe; }
        
        .form-container { background: #161616; padding: 30px; border-radius: 12px; max-width: 650px; margin: 0 auto 30px auto; box-