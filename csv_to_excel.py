import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def converteste_csv_in_excel(nume_fisier_csv):
    if not os.path.exists(nume_fisier_csv):
        print(f"❌ Eroare: Fișierul '{nume_fisier_csv}' nu a fost găsit!")
        return

    nume_fisier_excel = nume_fisier_csv.replace('.csv', '.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport Trafic"

    # Deschidem și citim CSV-ul
    with open(nume_fisier_csv, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        data = list(reader)

    if not data:
        print("❌ Fișierul CSV este gol.")
        return

    # --- DEFINIRE STILURI EXCEL ---
    # Cap de tabel (Albastru)
    header_fill = PatternFill(start_color="4facfe", end_color="4facfe", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Aliniere și margini (Borders)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Culori text pentru IN/OUT și rândul de Total
    in_font = Font(color="008000", bold=True)    # Verde
    out_font = Font(color="FF0000", bold=True)   # Roșu
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid") # Galben

    # --- SCRIEREA DATELOR ȘI FORMATATREA ---
    for row_idx, row in enumerate(data, start=1):
        ws.append(row)
        
        # Formatare pentru Capul de Tabel (primul rând)
        if row_idx == 1:
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
        else:
            if not row: # Trecem peste rândurile goale
                continue
                
            is_total_row = "TOTAL GENERAL" in row

            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = center_align
                cell.border = thin_border
                
                # Formatare specifică pentru rândul de TOTAL
                if is_total_row:
                    cell.font = total_font
                    cell.fill = total_fill
                else:
                    # Formatare condiționată pentru valorile numerice IN / OUT
                    # Coloana 4 = IN, Coloana 5 = OUT
                    if col_idx == 4 and str(cell.value).isdigit():
                        cell.font = in_font
                    elif col_idx == 5 and str(cell.value).isdigit():
                        cell.font = out_font

    # --- AJUSTAREA LĂȚIMII COLOANELOR ---
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 4 # Adăugăm padding

    # Salvăm fișierul
    wb.save(nume_fisier_excel)
    print(f"✅ Succes! Excelul a fost creat și aranjat frumos aici: {nume_fisier_excel}")

if __name__ == "__main__":
    print("=== CONVERTOR CSV ÎN EXCEL ===")
    fisier = input("Introdu numele fișierului CSV (ex: istoric_2026-02-20.csv): ")
    
    if not fisier.endswith('.csv'):
        fisier += '.csv'
        
    converteste_csv_in_excel(fisier)